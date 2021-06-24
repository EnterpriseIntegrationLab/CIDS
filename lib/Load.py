# Load CIDS in JSON format code for CA Repository

# Creator: Mark S. Fox, msf@eil.utoronto.ca
# Date: 26 January 2021
# Copyright 2021 Mark S. Fox
# Available under: http://creativecommons.org/licenses/by/3.0/


import datetime
from owlready2 import *
import types			# for dynamic class creation

from flask import Flask, render_template, request, session, redirect, url_for, g, flash

# Repository imports
import config
import Util
import Analysis

def requestLoad() :
	
	# kick user out if they are not a superuser - should not happen as they would not get access to the registerorganization page
	if (config.user.userType != cidsrep.superuser) and (config.user.userType != cidsrep.admin) :
		return(render_template('main.html', message="You do not have access rights for loading data."))
		
	return(render_template('loadJsonld.html', path="http://localhost:5000/LoadJSONLD"))


def load() :
	
	if request.method != 'POST': return(render_template('main.html', message="Incorrect post method."))
	if not config.organization : return(render_template('main.html', message="No organization selected."))
	
	# Part 1: save the file to the subdirectory for the organization
	#	* file contents cannot be binary!!! ascii string
    # check if the post request has the file part
	if not 'file'  in request.files :
		flash('No file part')
		return(redirect(request.url))

	file = request.files['file']
	# if user does not select file
	if file.filename == '':
		flash('No selected file')
		return(redirect(request.url))
	if not (file and allowed_file(file.filename)): render_template('main.html', message="File extension not allowed.")
    
    # store the file in the organization's upload directory
	filename = secure_filename(file.filename)
	print("Filename: ", filename)
	uploadDirectory = UPLOAD_FOLDER + '/' + config.organization.hasID.hasIdentifier
	if not os.path.exists(uploadDirectory): os.makedirs(uploadDirectory)
	path = os.path.join(uploadDirectory, filename)
	file.save(path)
    
    # Part 2: convert json-ld into n-triple and then load into OWLReady2
    # Next is to read and convert the file using rdflib serialization
	g = Graph()
	g.parse(path, format="json-ld")
	nt = g.serialize(format="nt")
	nt = nt.decode('ascii')
	ntpath, ext = path.split(".")
	ntpath = ntpath + ".owl"		# have to save with owl extension for owlready2 to compile entities properly
	with open(ntpath, 'w') as f: 
		f.write(nt)
		f.close()
	uploadns = cidsServer.get_ontology(ntpath) # now load the indiv into owlready2
	uploadns.load()
	
	# get ids of and log each one by re-reading the json-ld file and pulling ids
	with open(path, "r") as f: js = json.load(f)
	if type(js) != list : js = [js]
	for ind in js :
		id = ind["@id"]
		print("Loaded ", id)
		if id :
			idp = cidsServer.search_one(iri=id)
			if not idp : 
				print("Can't find ", id)
			else :
				logIndividual("Upload", idp)
	return(render_template('main.html', message="Upload complete: " + filename))
    

def loadNtriples(js) :
	graph = cidsServer.as_rdflib_graph()
	jsonDict = json.loads(js)
	id = jsonDict["@id"]
	typ = jsonDict["@type"]
	ind = cadr.search_one(iri=id)
	if ind :
		# check that the type is consistent, if not generate an errot
		ns, entity = rdflib.namespace.split_uri(ind)
		# delete entity and replace with new data
		logIndividual("Load JSONLD - delete", ind) # log it so that it can be retrieved if need
		delete_entity(ind)

	# create individual of the correct type
	type_ns, type_entity = rdflib.namespace.split_uri(typ)
	ind_ns, ind_entity = rdflib.namespace.split_uri(ind)
	tns = rdflib.Namespace(type_ns)
	ins = rdflib.Namespace(ind_ns)
	with cadr :
		graph.add((ins[ind_entity], RDF.type, tns[type_entity]))
	
	# go through keys and produce corresponding attribute
	# ADD HANDLING OF VALUE THAT IS A LIST
	for key in jsonDict :
		if (key != "@id") and (key != "@type") and (key != "@context") :
			pns, pentity = rdflib.namespace.split_uri(key)
			rdflib_pns = rdflib.Namespace(pns)
			values =  jsonDict[key] if type(jsonDict[key]) is list else [jsonDict[key]]
			vns, ventity = rdflib.namespace.split_uri(jsonDict[key])
			for value in values :
				rdflib_vns = rdflib.Namespace(vns)
				with cadr:
					graph.add((ins[ind_entity], rdflib_pns[pentity], rdflib_vns(ventity)))


#------------------- Load UN SDGs ------------------------
from openpyxl import load_workbook

def loadUNSDG(fname) :
	og = config.cids.StandardsOrganization(namespace=config.cadr)
	config.repository.hasOrganization.append(og)
	og.hasLegalName = "United Nations"
	og.hasID = config.org.OrganizationID(namespace=config.cadr, hasIdentifier="UN", forOrganization = og)
	og.hasDescription = "Global indicator framework for the Sustainable Development Goals and targets of the 2030 Agenda for Sustainable Development"
	og.hasIndicator = []
	og.hasOutcome = []
	og.hasCharacteristic = []
	
	# load UNSDGs
	workbook = load_workbook(filename=fname, read_only=True, data_only=True)
	sheet = workbook.active
	
	# cycle through each row and construct both Outcome and Indicator
	prevGoal = 0
	prevTarget = ""
	out = None
	
	for row in range(4,277) :
		goal = sheet.cell(row=row, column=1).value
		if not goal : continue		# skip heading line
		
		goal = int(goal)
		if prevGoal < goal :
			print("Starting Goal: ", goal)
			prevGoal = goal
			goalText = sheet.cell(row=row, column=4).value
			with config.cadr :
				outName = "UNSDGGoal" + str(goal)
				outcomeClass = types.new_class(outName, (config.cids.Outcome,))
			outcomeClass.comment = goalText
			outcomeClass.label = "UNSDG Goal " + str(goal)
			continue
		
		target = sheet.cell(row=row, column=2).value
		if target != prevTarget	:      # create new outcome
			prevTarget = target
			outcomeText = sheet.cell(row=row, column=4).value
			out = outcomeClass(namespace=config.cadr)
			out.hasDescription =  outcomeText + " [" + goalText + "]"
			out.hasIdentifier = target
			out.definedBy = og
			out.hasIndicator = []
			og.hasOutcome.append(out)
		
		indicatorText = sheet.cell(row=row, column=5).value
		code = sheet.cell(row=row, column=6).value
		ind = config.cids.Indicator(namespace=config.cadr)
		ind.hasName = code
		ind.hasDescription = indicatorText
		ind.definedBy = og
		ind.forOutcome = out
		og.hasIndicator.append(ind)
		out.hasIndicator.append(ind)
		

def loadIRIS(fname) :
	og = config.cids.StandardsOrganization(namespace=config.cadr)
	config.repository.hasOrganization.append(og)
	og.hasLegalName = "IRIS"
	og.hasID = config.org.OrganizationID(namespace=config.cadr, hasIdentifier="IRIS", forOrganization = og)
	og.hasDescription = "IRIS+ is the generally accepted system for measuring, managing, and optimizing impact."
	og.hasIndicator = []
	og.hasOutcome = []
	og.hasCharacteristic = []
	
	# load IRIS Indicators
	workbook = load_workbook(filename=fname, read_only=True, data_only=True)
	sheet = workbook.active
	
	# cycle through each row and construct both Outcome and Indicator	
	count = 0
	inds = og.hasIndicator
	for row in range(2,644) :
		count += 1
		code = sheet.cell(row=row, column=1).value
		indicatorName = sheet.cell(row=row, column=2).value
		indicatorDefinition = sheet.cell(row=row, column=3).value
		indicatorDescription = sheet.cell(row=row, column=6).value
	
		ind = config.cids.Indicator(namespace=config.cadr)
		ind.hasIdentifier = code
		ind.hasName = indicatorName
		ind.hasDescription = indicatorDescription
		ind.definedBy = og
		og.hasIndicator.append(ind)
		if count % 100 == 0 : print(count, " indicators created: ", indicatorName)
		
		


#------------------- Stakeholder characteristics loading ------------------------

# loads:
# organization source of characteristics - must already exist before loading
# loads characteristic name, description, id, type

def loadSC() :
	pass
